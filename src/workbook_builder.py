"""可追溯的 Excel 产物生成器。

保留现有模板中的已核对内容，只对 Version 4 要求做确定性增补。所有执行结果字段
保持“待执行/未提供”，不生成函证、盘点、错报或审计结论。
"""

from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.formatting.rule import DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.properties import CalcProperties

from .config import (
    CHECKLIST_OUTPUT_NAME,
    OUTPUT_DIR,
    PLAN_OUTPUT_NAME,
    ROOT,
    WORKBOOK_VERIFICATION_OUTPUT_NAME,
)
from .risk_engine import evaluate_risks


INK = "2D2D2D"
MUTED = "6D6E71"
ORANGE = "EB8C00"
RED = "E0301E"
DARK_RED = "A32020"
PALE = "F5F5F5"
LINE = "D9D9D9"
GREEN = "DDEBF7"
WHITE = "FFFFFF"
THIN = Side(style="thin", color=LINE)


# 高优先级只保留首周必须取得、会阻塞风险评估或不可逆外部程序的资料。
HIGH_PRIORITY_ITEMS = {
    "关联方及关联交易清单", "总账、明细账、科目余额表", "试算平衡表及财务报表", "银行账户清单",
    "销售收入明细", "年末前后销售清单", "前十大客户分析",
    "应收账款明细", "应收账款账龄表", "坏账准备计算表", "逾期款项清单", "期后回款", "函证联系人信息",
    "存货明细", "仓库清单", "盘点计划", "库龄分析", "存货跌价准备计算表", "委托加工物资明细",
    "年末前后入出库明细", "生产成本计算表",
    "采购明细", "制造费用分配表", "单位成本计算表", "成本结转表",
    "银行对账单", "银行余额调节表", "借款合同", "银行函证信息", "未来12个月资金预算",
    "预付款项明细", "其他流动资产明细", "主要预付款合同与付款", "预付款期后到货结算",
    "应付账款及采购负债明细", "供应商对账及函证信息", "年末前后采购入库清单",
    "承诺及或有事项清单",
}

LOW_PRIORITY_ITEMS = {
    "期后事项清单", "客户对账单", "销售价格表", "盘点表及差异表", "期后销售价格",
    "政府补助资料", "长期待摊费用明细", "无形资产明细", "员工花名册与薪酬汇总",
    "费用合同与抽样支持资料", "财务报表附注及披露清单", "关联方披露勾稽表", "报表列报完整性核对表",
}

EXTERNAL_PROCEDURE_ITEMS = {
    "银行账户清单", "函证联系人信息", "仓库清单", "盘点计划", "委托加工物资明细",
    "银行函证信息", "供应商对账及函证信息",
}

FOUNDATION_ITEMS = {
    "关联方及关联交易清单", "总账、明细账、科目余额表", "试算平衡表及财务报表",
}


def _priority_and_reason(material_name: str) -> tuple[str, str]:
    if material_name in HIGH_PRIORITY_ITEMS:
        if material_name in EXTERNAL_PROCEDURE_ITEMS:
            return "高", "首周取得：用于函证、盘点等不可逆外部程序的范围或地址核验"
        if material_name in FOUNDATION_ITEMS:
            return "高", "首周取得：影响整体风险评估、报表勾稽和后续选样"
        return "高", "首周取得：构成重点项目完整总体、关键估计或截止测试基础"
    if material_name in LOW_PRIORITY_ITEMS:
        return "低", "收尾或定向补充：用于披露、补充验证或执行结果形成后的跟进"
    return "中", "实质性程序前取得：用于流程了解、样本支持或一般报表项目测试"


def _apply_checklist_priority_policy(wb: Workbook) -> dict[str, int]:
    counts = {"高": 0, "中": 0, "低": 0}
    for ws in wb.worksheets:
        if ws.title == "总览":
            continue
        ws["L1"] = "优先级依据/备注"
        for row in range(2, ws.max_row + 1):
            material_name = str(ws.cell(row, 3).value or "")
            priority, reason = _priority_and_reason(material_name)
            ws.cell(row, 8, priority)
            ws.cell(row, 12, reason)
            planned = ws.cell(row, 9).value
            planned_date = planned.date() if isinstance(planned, datetime) else planned
            if priority == "高" and (not isinstance(planned_date, date) or planned_date > date(2023, 1, 6)):
                ws.cell(row, 9, date(2023, 1, 6))
            elif priority == "低" and (not isinstance(planned_date, date) or planned_date < date(2023, 1, 20)):
                ws.cell(row, 9, date(2023, 1, 20))
            ws.cell(row, 9).number_format = "yyyy-mm-dd"
            counts[priority] += 1
        _set_widths(ws, [8, 18, 30, 44, 16, 16, 20, 10, 16, 16, 12, 38])
    return counts


def _style_header(ws, row: int, start: int, end: int) -> None:
    for cell in ws.iter_cols(min_col=start, max_col=end, min_row=row, max_row=row):
        c = cell[0]
        c.fill = PatternFill("solid", fgColor=ORANGE)
        c.font = Font(bold=True, color=WHITE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(bottom=Side(style="medium", color=DARK_RED))
    ws.row_dimensions[row].height = 30


def _style_body(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.font = Font(name="Arial", size=10, color=INK)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(bottom=THIN)


def _set_widths(ws, widths: list[float]) -> None:
    from openpyxl.utils import get_column_letter

    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def _replace_table(ws, name: str, ref: str) -> None:
    for existing in list(ws.tables.values()):
        del ws.tables[existing.name]
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(table)


def _enable_formula_recalculation(wb: Workbook) -> None:
    wb.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)


NEW_CHECKLIST_MODULES: dict[str, list[tuple[str, str, str, str, str, str, str]]] = {
    "其他报表项目": [
        ("应付账款及采购负债明细", "供应商、发票、账龄、未结算入库及期后付款", "2022/12/31", "Excel", "采购部/财务部", "高", "2023-01-05"),
        ("供应商对账及函证信息", "主要及异常供应商对账单、独立地址来源", "2022/12/31", "Excel/PDF", "采购部", "高", "2023-01-06"),
        ("年末前后采购入库清单", "2022-12-15 至 2023-01-15 全量采购与入库", "年末前后", "Excel", "采购部/仓库", "高", "2023-01-05"),
        ("固定资产卡片及变动表", "类别、原值、折旧、地点、责任人及本年增减", "2022", "Excel", "财务部/设备部", "中", "2023-01-09"),
        ("固定资产采购验收与处置资料", "合同、发票、验收、付款、报废及审批", "2022", "PDF", "设备部/财务部", "中", "2023-01-10"),
        ("在建工程明细与工程合同", "项目进度、预算、付款、验收及转固判断", "2022", "Excel/PDF", "工程部/财务部", "中", "2023-01-10"),
        ("长期待摊费用明细", "项目、原始支出、摊销期、摊销额及受益期间", "2022", "Excel/PDF", "财务部", "中", "2023-01-10"),
        ("租赁合同完整清单", "租赁标的、期限、付款、续租及终止选择权", "2022及期后", "Excel/PDF", "行政部/财务部", "中", "2023-01-09"),
        ("使用权资产及租赁负债测算", "折现率、付款计划、变更、利息及折旧", "2022", "Excel", "财务部", "中", "2023-01-10"),
        ("无形资产明细", "类别、取得方式、使用寿命、摊销及减值迹象", "2022", "Excel/PDF", "财务部/法务", "中", "2023-01-10"),
        ("应付职工薪酬明细", "工资、奖金、社保、公积金及期后支付", "2022及期后", "Excel", "人力资源部/财务部", "中", "2023-01-09"),
        ("员工花名册与薪酬汇总", "人员、部门、入离职、工资构成及个税", "2022", "Excel", "人力资源部", "中", "2023-01-09"),
        ("其他资产负债项目明细", "逐项构成、账龄、对手方、原始凭证和期后结转", "2022/12/31", "Excel/PDF", "财务部", "中", "2023-01-10"),
    ],
    "费用税务与披露": [
        ("税费明细及纳税申报表", "增值税、附加税、企业所得税及申报勾稽", "2022", "Excel/PDF", "税务/财务部", "高", "2023-01-09"),
        ("所得税计算表", "当期所得税、递延所得税及纳税调整", "2022", "Excel", "税务/财务部", "高", "2023-01-10"),
        ("税务稽查及争议资料", "检查通知、补税、处罚、沟通及未决事项", "2022及期后", "PDF", "税务/法务", "高", "2023-01-10"),
        ("销售费用明细", "按月份、部门、项目及对手方的全年明细", "2022", "Excel", "财务部", "中", "2023-01-09"),
        ("管理费用明细", "按月份、部门、项目及对手方的全年明细", "2022", "Excel", "财务部", "中", "2023-01-09"),
        ("研发费用明细", "项目、人员、材料、折旧及资本化判断", "2022", "Excel", "研发部/财务部", "中", "2023-01-09"),
        ("财务费用明细", "利息、汇兑损益、手续费及勾稽资料", "2022", "Excel", "财务部", "中", "2023-01-09"),
        ("费用合同与抽样支持资料", "合同、发票、审批、付款及受益期间", "2022及期后", "PDF", "各部门/财务部", "中", "2023-01-11"),
        ("现金流量表及编制底稿", "主表、补充资料、非现金事项及勾稽", "2022", "Excel", "财务部", "高", "2023-01-10"),
        ("财务报表附注及披露清单", "会计政策、报表项目、风险及关联方披露", "2022", "Excel/Word", "财务部", "高", "2023-01-10"),
        ("承诺及或有事项清单", "资本承诺、担保、诉讼、索赔及预计负债", "2022及期后", "Excel/PDF", "法务/财务部", "高", "2023-01-10"),
        ("关联方披露勾稽表", "关联关系、交易、余额、定价及审批", "2022", "Excel", "董事会办公室/财务部", "高", "2023-01-10"),
        ("报表列报完整性核对表", "报表、附注、比较数据和勾稽关系", "2022", "Excel", "财务部", "高", "2023-01-12"),
    ],
}


def _build_checklist_overview(wb: Workbook) -> None:
    if "总览" in wb.sheetnames:
        del wb["总览"]
    ws = wb.create_sheet("总览", 0)
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:J1")
    ws["A1"] = "A 公司 2022 年度审计资料清单"
    ws["A1"].fill = PatternFill("solid", fgColor=INK)
    ws["A1"].font = Font(size=18, bold=True, color=WHITE)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 36
    ws.merge_cells("A2:J2")
    ws["A2"] = "模拟审计项目｜未列为高风险的重大报表项目仍执行基础程序｜状态与日期均为可编辑字段"
    ws["A2"].fill = PatternFill("solid", fgColor=PALE)
    ws["A2"].font = Font(color=MUTED, italic=True)
    ws["A2"].alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28
    # 资料清单沿用同一模拟状态日，避免历史计划被 TODAY() 误标为逾期。
    ws["L1"] = "项目参数"
    ws["L2"] = "项目模式"; ws["M2"] = "计划"
    ws["L3"] = "项目状态日"; ws["M3"] = date(2023, 1, 10)
    ws["M3"].number_format = "yyyy-mm-dd"
    ws.column_dimensions["L"].hidden = True
    ws.column_dimensions["M"].hidden = True
    headers = ["模块", "资料总数", "高优先级", "中优先级", "低优先级", "已提供", "待补充", "未提供", "完成率", "负责人提示"]
    ws.append([])
    ws.append(headers)
    modules = [name for name in wb.sheetnames if name != "总览"]
    for index, name in enumerate(modules, start=6):
        last = wb[name].max_row
        ws.append([
            name, f"=COUNTA('{name}'!$A$2:$A${last})", f'=COUNTIF(\'{name}\'!$H$2:$H${last},"高")',
            f'=COUNTIF(\'{name}\'!$H$2:$H${last},"中")', f'=COUNTIF(\'{name}\'!$H$2:$H${last},"低")',
            f'=COUNTIF(\'{name}\'!$K$2:$K${last},"已提供")', f'=COUNTIF(\'{name}\'!$K$2:$K${last},"待补充")',
            f'=COUNTIF(\'{name}\'!$K$2:$K${last},"未提供")', f"=IF(B{index}=0,0,F{index}/B{index})", "对应模块负责人跟进",
        ])
    total_row = ws.max_row + 1
    ws.append(["合计", f"=SUM(B6:B{total_row-1})", f"=SUM(C6:C{total_row-1})", f"=SUM(D6:D{total_row-1})", f"=SUM(E6:E{total_row-1})", f"=SUM(F6:F{total_row-1})", f"=SUM(G6:G{total_row-1})", f"=SUM(H6:H{total_row-1})", f"=IF(B{total_row}=0,0,F{total_row}/B{total_row})", "每日更新"])
    _style_header(ws, 5, 1, 10)
    _style_body(ws, 6, total_row, 1, 10)
    for c in ws[total_row]:
        c.fill = PatternFill("solid", fgColor=PALE)
        c.font = Font(bold=True, color=INK)
    for row in range(6, total_row + 1):
        ws.cell(row, 9).number_format = "0.0%"
    ws.conditional_formatting.add(
        f"I6:I{total_row}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=ORANGE),
    )
    _set_widths(ws, [20, 12, 12, 12, 12, 12, 12, 12, 12, 26])
    ws.freeze_panes = "A6"
    _replace_table(ws, "ChecklistOverview", f"A5:J{total_row}")


def _normalize_checklist_overdue_rule(ws) -> None:
    """Replace template TODAY() rules with the shared historical project parameters."""
    conditional_formatting = ws.conditional_formatting
    for key, rules in list(conditional_formatting._cf_rules.items()):
        retained = []
        for rule in rules:
            formulas = rule.formula or []
            is_overdue_rule = any('$K2<>"已提供"' in formula and "$I2<" in formula for formula in formulas)
            if not is_overdue_rule:
                retained.append(rule)
        if retained:
            conditional_formatting._cf_rules[key] = retained
        else:
            del conditional_formatting._cf_rules[key]
    conditional_formatting.add(
        f"A2:L{ws.max_row}",
        FormulaRule(
            formula=['AND(\'总览\'!$M$2="执行",$K2<>"已提供",$I2<\'总览\'!$M$3)'],
            fill=PatternFill("solid", fgColor="FDE9E7"),
        ),
    )


def build_checklist_workbook() -> dict[str, object]:
    template = ROOT / "templates" / "audit_checklist_template.xlsx"
    wb = load_workbook(template)
    for module, items in NEW_CHECKLIST_MODULES.items():
        if module in wb.sheetnames:
            del wb[module]
        ws = wb.create_sheet(module)
        ws.sheet_view.showGridLines = False
        headers = ["序号", "模块", "资料名称", "资料说明", "对应期间", "格式要求", "责任部门", "优先级", "计划提供日期", "实际提供日期", "状态", "备注"]
        ws.append(headers)
        for idx, item in enumerate(items, start=1):
            ws.append([idx, module, item[0], item[1], item[2], item[3], item[4], item[5], date.fromisoformat(item[6]), None, "未提供", ""])
        _style_header(ws, 1, 1, 12)
        _style_body(ws, 2, ws.max_row, 1, 12)
        _set_widths(ws, [8, 18, 30, 44, 16, 16, 20, 10, 16, 16, 12, 26])
        ws.freeze_panes = "A2"
        for row in range(2, ws.max_row + 1):
            ws.cell(row, 9).number_format = "yyyy-mm-dd"
            ws.cell(row, 10).number_format = "yyyy-mm-dd"
        priority = DataValidation(type="list", formula1='"高,中,低"')
        status = DataValidation(type="list", formula1='"未提供,已提供,待补充"')
        ws.add_data_validation(priority); priority.add(f"H2:H{ws.max_row}")
        ws.add_data_validation(status); status.add(f"K2:K{ws.max_row}")
        ws.conditional_formatting.add(f"K2:K{ws.max_row}", FormulaRule(formula=['$K2="已提供"'], fill=PatternFill("solid", fgColor="E2F0D9")))
        ws.conditional_formatting.add(f"K2:K{ws.max_row}", FormulaRule(formula=['$K2="待补充"'], fill=PatternFill("solid", fgColor="FFF2CC")))
        _replace_table(ws, f"Checklist_{len(wb.sheetnames)}", f"A1:L{ws.max_row}")
    priority_counts = _apply_checklist_priority_policy(wb)
    _build_checklist_overview(wb)
    for ws in wb.worksheets:
        if ws.title != "总览":
            _normalize_checklist_overdue_rule(ws)
    target = OUTPUT_DIR / CHECKLIST_OUTPUT_NAME
    target.parent.mkdir(parents=True, exist_ok=True)
    _enable_formula_recalculation(wb)
    wb.save(target)
    result = {
        "file": target.name,
        "sheet_names": wb.sheetnames,
        "sheet_count": len(wb.sheetnames),
        "new_modules": {name: len(items) for name, items in NEW_CHECKLIST_MODULES.items()},
        "total_requests": sum(wb[name].max_row - 1 for name in wb.sheetnames if name != "总览"),
        "priority_counts": priority_counts,
        "high_priority_ratio": priority_counts["高"] / sum(priority_counts.values()),
    }
    return result


ADDITIONAL_PROCEDURES = [
    ("AP-01", "应付账款与采购负债", "一般", "采购负债可能漏记或跨期", "完整性；截止；计价", "供应商对账/函证；检查期后付款、未结算入库及年末前后入库发票", "高级审计员B", "项目经理", "2023-01-16", "2023-01-20", "采购流程测试"),
    ("FA-01", "固定资产与在建工程", "一般", "增减变动、转固和折旧可能不准确", "存在；权利和义务；计价；列报", "核对卡片；抽查购置验收与处置；现场观察；重算折旧并复核转固时点", "审计员A", "项目经理", "2023-01-16", "2023-01-20", "固定资产卡片"),
    ("LTD-01", "长期待摊费用", "一般", "摊销期间及费用资本化可能不恰当", "发生；计价；分类", "检查原始支出和受益期；重算摊销；评价资本化与费用化分类", "审计员A", "项目经理", "2023-01-18", "2023-01-20", "长期待摊明细"),
    ("LEASE-01", "租赁", "一般", "租赁完整性、折现率和计量涉及判断", "完整性；计价；列报", "检查合同完整清单；复核租赁识别、折现率和付款；重算使用权资产与租赁负债", "审计员A", "项目经理", "2023-01-16", "2023-01-20", "租赁合同及测算"),
    ("PAY-01", "职工薪酬", "一般", "人员、薪酬计提及跨期可能不准确", "发生；完整性；截止；计价", "花名册与工资表勾稽；抽查审批和支付；重算奖金社保并检查期后支付", "审计员B", "高级审计员B", "2023-01-16", "2023-01-20", "薪酬明细与花名册"),
    ("TAX-01", "税费及所得税", "一般", "申报勾稽、纳税调整和递延所得税可能不准确", "完整性；计价；列报", "税费明细与申报表勾稽；重算所得税；复核递延所得税及未决税务事项", "审计员A", "项目经理", "2023-01-16", "2023-01-23", "税务申报与所得税测算"),
    ("EXP-01", "期间费用", "一般", "费用可能跨期、分类错误或缺少商业实质", "发生；完整性；截止；分类", "趋势分析；大额异常及期末样本抽查；检查合同发票审批付款及受益期", "审计员B", "高级审计员B", "2023-01-16", "2023-01-20", "费用全量明细"),
    ("IA-01", "无形资产", "一般", "权属、摊销年限及减值判断可能不恰当", "权利和义务；计价；列报", "检查权属和取得资料；重算摊销；评价使用寿命及减值迹象", "审计员A", "项目经理", "2023-01-18", "2023-01-20", "无形资产明细"),
    ("CF-01", "现金流量表", "一般", "现金流分类和补充资料勾稽可能不准确", "准确性；分类；列报", "复核编制底稿；与总账及报表勾稽；抽查重大非现金及分类事项", "审计员B", "项目经理", "2023-01-20", "2023-01-24", "现金流量表底稿"),
    ("DISC-01", "附注与或有事项", "一般", "披露可能不完整或与主表不一致", "完整性；列报", "使用披露清单；勾稽主表附注；检查承诺、诉讼、担保、关联方和期后事项", "项目经理", "项目合伙人", "2023-01-20", "2023-01-25", "附注及法务资料"),
]


CONTROL_ROWS = [
    ("销售与收款", "订单、发货或收入确认未经适当审批", "订单信用审批；发货与开票匹配；月末截止复核", "逐笔/每月", "销售及财务负责人", "待了解", "待决定", "询问、观察、穿行及抽样检查", "暂不预设依赖；不能依赖时增加收入抽凭与截止测试"),
    ("采购与付款", "未授权采购、重复付款或负债漏记", "供应商准入；三单匹配；付款双人审批", "逐笔", "采购及财务负责人", "待了解", "待决定", "询问、穿行、检查审批与系统日志", "不能依赖时扩大采购截止和未记录负债测试"),
    ("生产与存货", "领料、产量、成本归集或盘点记录不完整", "工单领料审批；成本月结复核；定期盘点", "逐批/每月", "生产及仓库负责人", "待了解", "待决定", "穿行、观察、检查盘点与复核证据", "不能依赖时扩大抽盘、成本重算与截止测试"),
    ("资金", "未授权付款、账户遗漏或资金被占用", "网银分权；银行余额调节；账户清单定期复核", "逐笔/每月", "资金负责人", "待了解", "待决定", "观察权限、检查调节表和审批记录", "不能依赖时增加银行函证及流水测试"),
    ("财务结账", "手工分录绕过控制或报表勾稽错误", "分录审批；结账清单；报表附注复核", "每月/每年", "财务负责人", "待了解", "待决定", "检查分录权限、结账清单和复核痕迹", "不能依赖时扩大分录测试和实质性分析"),
    ("主数据和系统权限", "不当权限或主数据修改影响交易完整性", "账号审批；定期权限复核；主数据变更日志", "按需/季度", "信息技术负责人", "待了解", "待决定", "检查用户清单、权限矩阵、变更日志及离职账号", "IT 一般控制不足时不依赖自动化控制并增加实质性程序"),
]


def _replace_sheet(wb: Workbook, title: str, index: int | None = None):
    if title in wb.sheetnames:
        del wb[title]
    return wb.create_sheet(title, index)


def _add_materiality_sheet(wb: Workbook) -> None:
    ws = _replace_sheet(wb, "重要性与抽样")
    ws.sheet_view.showGridLines = False
    ws.append(["重要性与抽样决策模板", "内容/公式", "责任人", "复核人", "状态/说明"])
    rows = [
        ("重要性基准", "候选基准：税前利润、营业收入或总资产；正式选择待项目经理判断", "项目经理", "项目合伙人", "案例报表可作候选基准比较，正式选择待审批"),
        ("基准金额", "待取得经核对的试算平衡表后填写", "项目经理", "项目合伙人", "待确定"),
        ("选用比例", "待结合使用者、波动和基准稳定性判断", "项目经理", "项目合伙人", "待确定"),
        ("财务报表整体重要性", '=IF(AND(ISNUMBER(B3),ISNUMBER(B4)),B3*B4,"待项目经理确定")', "项目经理", "项目合伙人", "公式计算"),
        ("实际执行重要性", '=IF(AND(ISNUMBER(B5),ISNUMBER(B7)),B5*B7,"待项目经理确定")', "项目经理", "项目合伙人", "比例在 B7 填写"),
        ("实际执行重要性比例", "待结合错报历史、控制风险和汇总风险判断", "项目经理", "项目合伙人", "待确定"),
        ("明显微小错报临界值", '=IF(AND(ISNUMBER(B5),ISNUMBER(B9)),B5*B9,"待项目经理确定")', "项目经理", "项目合伙人", "比例在 B9 填写"),
        ("明显微小错报比例", "待结合定性因素判断", "项目经理", "项目合伙人", "待确定"),
        ("定性重要性因素", "关联方、管理层舞弊、监管事项、契约条款、敏感披露", "项目经理", "项目合伙人", "金额以下事项仍可能重大"),
    ]
    for row in rows:
        ws.append(row)
    start = ws.max_row + 2
    ws.cell(start, 1, "抽样总体"); ws.cell(start, 2, "选样方法"); ws.cell(start, 3, "样本量依据"); ws.cell(start, 4, "与风险/重要性的联系"); ws.cell(start, 5, "状态")
    samples = [
        ("收入", "大额全查 + 年末/手工/异常毛利定向 + 其余随机或系统抽样", "总体规模、风险等级、执行重要性、控制依赖", "高风险；侧重发生与截止", "待重要性确定后计算"),
        ("采购与应付", "大额及未结算入库定向 + 年末前后系统抽样", "完整性风险、总体规模、控制依赖", "侧重完整性与截止", "待重要性确定后计算"),
        ("应收账款", "重大/异常余额全选 + 其余分层抽样函证", "余额覆盖率、账龄、争议、执行重要性", "高风险；存在与计价", "待明细取得后计算"),
        ("存货", "重大库位/品类 + 呆滞异常 + 其余随机双向抽盘", "金额、库龄、地点、控制及执行重要性", "高风险；存在与完整性", "待盘点表取得后计算"),
        ("期间费用", "大额异常、关联方、期末样本 + 随机抽样", "科目波动、执行重要性和舞弊因素", "一般风险基础程序", "待总账取得后计算"),
    ]
    for row in samples:
        ws.append(row)
    _style_header(ws, 1, 1, 5); _style_body(ws, 2, 10, 1, 5)
    _style_header(ws, start, 1, 5); _style_body(ws, start + 1, ws.max_row, 1, 5)
    _set_widths(ws, [26, 48, 40, 36, 28]); ws.freeze_panes = "A2"


def _add_risk_sheet(wb: Workbook) -> None:
    ws = _replace_sheet(wb, "风险登记表")
    headers = ["风险 ID", "风险项目", "金额重大性", "同比波动", "管理层判断程度", "舞弊可能性", "交易复杂度", "证据获取难度", "平均分", "计算等级", "人工调整等级", "调整原因", "最终等级", "评分依据"]
    ws.append(headers)
    score_names = headers[2:8]
    for idx, risk in enumerate(evaluate_risks(), start=2):
        reason = "；".join(risk.reasons)
        ws.append([
            risk.risk_id,
            risk.area,
            *[risk.scores[name] for name in score_names],
            f"=AVERAGE(C{idx}:H{idx})",
            f'=IF(I{idx}>=4,"高",IF(I{idx}>=3,"中高",IF(I{idx}>=2,"中","一般")))',
            "",
            "",
            f'=IF(K{idx}="",J{idx},K{idx})',
            reason,
        ])
    base_row = ws.max_row + 1
    ws.append([
        "R-BASE", "其他报表项目基础程序", None, None, None, None, None, None, None,
        "一般", "", "未识别为专项风险，仍执行基础程序", f'=IF(K{base_row}="",J{base_row},K{base_row})',
        "基础程序不代表零风险；如风险评估变化，应在本表调整并说明。",
    ])
    _style_header(ws, 1, 1, len(headers)); _style_body(ws, 2, ws.max_row, 1, len(headers))
    _set_widths(ws, [14, 26, 12, 12, 16, 12, 12, 14, 12, 12, 14, 34, 12, 60])
    level = DataValidation(type="list", formula1='"高,中高,中,一般"', allow_blank=True)
    ws.add_data_validation(level); level.add(f"K2:K{ws.max_row}")
    last_data_row = ws.max_row
    ws.freeze_panes = "C2"; _replace_table(ws, "RiskRegister", f"A1:N{last_data_row}")
    note_row = last_data_row + 2
    ws.cell(note_row, 1, "说明")
    ws.cell(note_row, 2, "评分仅用于资源排序，不替代职业判断；人工覆盖必须填写调整原因。")
    ws.merge_cells(start_row=note_row, start_column=2, end_row=note_row, end_column=14)


def _add_controls_sheet(wb: Workbook) -> None:
    ws = _replace_sheet(wb, "控制了解")
    headers = ["流程", "风险", "关键控制", "控制频率", "控制责任人", "设计有效性", "是否测试运行有效性", "测试方法", "结论/实质性应对"]
    ws.append(headers)
    for row in CONTROL_ROWS:
        ws.append(row)
    _style_header(ws, 1, 1, 9); _style_body(ws, 2, ws.max_row, 1, 9)
    _set_widths(ws, [18, 34, 44, 14, 20, 16, 20, 36, 52])
    design = DataValidation(type="list", formula1='"待了解,有效,无效,不适用"')
    test = DataValidation(type="list", formula1='"待决定,是,否"')
    ws.add_data_validation(design); design.add(f"F2:F{ws.max_row}")
    ws.add_data_validation(test); test.add(f"G2:G{ws.max_row}")
    ws.freeze_panes = "A2"; _replace_table(ws, "ControlMatrix", f"A1:I{ws.max_row}")


def _add_project_parameters(wb: Workbook) -> None:
    ws = _replace_sheet(wb, "项目参数", 0)
    ws.sheet_view.showGridLines = False
    ws.append(["项目参数", "参数值", "说明"])
    ws.append(["项目模式", "计划", "计划模式不计算逾期；切换为执行后按项目状态日计算"])
    ws.append(["项目状态日", date(2023, 1, 10), "模拟项目进度截止日，不使用现实 TODAY()"])
    ws.append(["汇报日期", date(2026, 8, 5), "Version 4 汇报日期"])
    ws.append(["模拟审计开始日", date(2022, 12, 27), "盘点准备前置开始日"])
    ws.append(["模拟审计结束日", date(2023, 1, 27), "收尾与独立复核完成日"])
    _style_header(ws, 1, 1, 3); _style_body(ws, 2, 6, 1, 3)
    _set_widths(ws, [24, 20, 66])
    for row in range(3, 7):
        ws.cell(row, 2).number_format = "yyyy-mm-dd"
    mode = DataValidation(type="list", formula1='"计划,执行"')
    ws.add_data_validation(mode); mode.add("B2")
    ws.freeze_panes = "A2"
    _replace_table(ws, "ProjectParameters", "A1:C6")


def _add_dashboard(wb: Workbook, last_proc_row: int) -> None:
    ws = _replace_sheet(wb, "项目管理看板", 0)
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1"); ws["A1"] = "项目管理看板（计划阶段）"
    ws["A1"].fill = PatternFill("solid", fgColor=INK); ws["A1"].font = Font(size=18, bold=True, color=WHITE)
    ws["A1"].alignment = Alignment(vertical="center"); ws.row_dimensions[1].height = 36
    ws.append(["指标", "公式/结果", "解释", "指标", "公式/结果", "解释"])
    metrics = [
        ("程序总数", f"=COUNTA('审计程序'!$A$2:$A${last_proc_row})", "覆盖重点与一般报表项目", "高风险程序", f'=COUNTIF(\'审计程序\'!$I$2:$I${last_proc_row},"高")', "风险等级由风险登记表同步"),
        ("已完成任务", f'=COUNTIF(\'审计程序\'!$R$2:$R${last_proc_row},"已完成")', "仅执行后更新", "已完成待复核", f'=COUNTIFS(\'审计程序\'!$R$2:$R${last_proc_row},"已完成",\'审计程序\'!$Z$2:$Z${last_proc_row},"<>已复核")', "未开始任务不计入"),
        ("受阻任务", f'=COUNTIF(\'审计程序\'!$Y$2:$Y${last_proc_row},"<>")', "按阻塞原因非空统计", "逾期任务", f'=COUNTIF(\'审计程序\'!$X$2:$X${last_proc_row},">0")', "按项目状态日计算"),
        ("资料完成率", f"=IFERROR('[{CHECKLIST_OUTPUT_NAME}]总览'!$I$15,0)", "从资料清单总览合计行读取", "函证回函率", '=IFERROR(COUNT(\'函证跟踪\'!$H$2:$H$21)/COUNT(\'函证跟踪\'!$F$2:$F$21),0)', "分母仅为实际首次发出"),
        ("盘点完成率", '=IFERROR(COUNTIF(\'盘点安排\'!$M$2:$M$13,"已完成")/COUNTIF(\'盘点安排\'!$B$2:$B$13,"<>"),0)', "分母仅为已安排地点", "自我复核冲突", f'=SUMPRODUCT(--(\'审计程序\'!$G$2:$G${last_proc_row}=\'审计程序\'!$H$2:$H${last_proc_row}))', "必须为 0"),
    ]
    for row in metrics:
        ws.append(row)
    _style_header(ws, 2, 1, 6); _style_body(ws, 3, ws.max_row, 1, 6)
    _set_widths(ws, [24, 18, 34, 24, 18, 34])


def build_audit_plan_workbook() -> dict[str, object]:
    template = ROOT / "templates" / "audit_plan_template.xlsx"
    wb = load_workbook(template)

    # 时间表：显式加入重要性任务，工期统一为工作日。
    timeline = wb["总体时间表"]
    timeline["A2"] = "报告期：2022 年｜以下为模拟年报审计时间轴，与汇报日期无关｜工期按工作日（NETWORKDAYS）计算｜复核关系不得取消"
    insert_row = 7
    timeline.insert_rows(insert_row)
    timeline.cell(insert_row, 1, "第1周")
    timeline.cell(insert_row, 2, "确定并复核财务报表整体重要性、实际执行重要性、明显微小错报临界值及抽样策略")
    timeline.cell(insert_row, 3, date(2023, 1, 5)); timeline.cell(insert_row, 4, date(2023, 1, 6))
    timeline.cell(insert_row, 6, "项目经理"); timeline.cell(insert_row, 7, "重要性与抽样决策记录"); timeline.cell(insert_row, 8, "未开始")
    for row in range(5, timeline.max_row + 1):
        timeline.cell(row, 5, f"=NETWORKDAYS(C{row},D{row})")
        timeline.cell(row, 3).number_format = timeline.cell(row, 4).number_format = "yyyy-mm-dd"
    _style_body(timeline, 5, timeline.max_row, 1, 8)
    _replace_table(timeline, "AuditTimeline", f"A4:H{timeline.max_row}")

    # 人员分工：项目合伙人与质量复核人分别列示，避免合并职责。
    roles = wb["人员分工"]
    roles.delete_rows(1, roles.max_row)
    roles.append(["角色", "主要职责", "主要模块", "重点复核事项", "协作关系"])
    role_rows = [
        ("项目合伙人", "对项目质量承担总体责任，审批重要性并复核重大判断与最终结论", "重大判断及整体层面", "重要性、持续经营、舞弊风险、重大差异和报告结论", "复核项目经理形成的判断，不参与同一事项的编制"),
        ("质量复核人（如适用）", "客观评价重大判断和拟出具报告，不代替项目合伙人职责", "项目质量复核", "重大判断、独立性、报告适当性及未决事项", "独立于项目组，不参与项目执行或决策形成"),
        ("项目经理", "总体计划、风险判断、团队管理、质量复核和管理层沟通", "全部模块", "关键判断、重大差异和项目结论", "向项目合伙人汇报并安排项目组任务"),
        ("高级审计员A", "收入、应收账款、客户函证和外销测试", "收入/应收", "函证样本、截止测试、坏账准备", "审计员B支持数据分析"),
        ("高级审计员B", "存货、成本、盘点和毛利率分析", "存货/成本", "监盘、跌价、成本重算", "审计员B支持数据整理"),
        ("审计员A", "资金、借款、费用、预付款和其他流动资产", "资金/其他", "银行函证、流动性、预付期后结转", "项目经理直接复核"),
        ("审计员B", "全量数据整理、分析程序、样本准备和底稿归档", "跨模块", "数据完整性、图表、索引", "两名高级审计员复核"),
    ]
    for role_row in role_rows:
        roles.append(role_row)
    _style_header(roles, 1, 1, 5); _style_body(roles, 2, roles.max_row, 1, 5)
    _replace_table(roles, "RolesTable", f"A1:E{roles.max_row}")

    proc = wb["审计程序"]
    old_headers = [proc.cell(1, col).value for col in range(1, proc.max_column + 1)]
    original_rows = [[proc.cell(row, col).value for col in range(1, proc.max_column + 1)] for row in range(2, proc.max_row + 1)]
    headers = ["工作编号", "风险 ID", "审计模块", "风险描述", "审计认定", "审计程序", "负责人", "复核人", "风险等级", "程序执行优先级", "前置任务", "计划开始", "计划完成", "实际完成", "预计工时", "实际工时", "完成百分比", "状态", "证据索引", "发现的问题", "结论", "依赖资料", "资料状态", "逾期工作日", "阻塞原因", "复核状态", "最终底稿索引"]
    proc.delete_rows(1, proc.max_row)
    proc.data_validations.dataValidation = []
    proc.conditional_formatting = ConditionalFormattingList()
    proc.append(headers)
    independent = {"GC-01", "J-01", "RP-01", "SE-01", "F-01"}
    def procedure_risk_id(code: str) -> str:
        if code.startswith("R-"):
            return "R-REV"
        if code.startswith("AR-"):
            return "R-AR"
        if code.startswith("I-"):
            return "R-INV"
        if code.startswith("C-"):
            return "R-COST"
        if code in {"P-01", "OCA-01"}:
            return "R-PREOCA"
        if code in {"B-01", "GC-01"}:
            return "R-DEBTGC"
        if code == "E-01":
            return "R-EQUITY"
        return "R-BASE"

    risk_last_row = len(evaluate_risks()) + 2
    for old in original_rows:
        code, area, risk, assertions, procedure, owner, reviewer, start, finish = old[:9]
        if code == "I-01":
            procedure = "参加年末盘点；账到物和物到账抽盘；检查委托加工存货；如存在异地存货，取得第三方证据"
        if code in independent or owner == reviewer:
            reviewer = "项目合伙人"
        proc.append([code, procedure_risk_id(code), area, risk, assertions, procedure, owner, reviewer, None, None, "风险评估/资料取得", start, finish, None, None, None, 0, "未开始", "", "", "待执行", "对应模块资料", "未提供", None, "", "未复核", ""])
    for code, area, level, risk, assertions, procedure, owner, reviewer, start, finish, dependency in ADDITIONAL_PROCEDURES:
        proc.append([code, "R-BASE", area, risk, assertions, procedure, owner, reviewer, None, None, "基础分析/资料取得", date.fromisoformat(start), date.fromisoformat(finish), None, None, None, 0, "未开始", "", "", "待执行", dependency, "未提供", None, "", "未复核", ""])
    for row in range(2, proc.max_row + 1):
        proc.cell(row, 9, f'=IFERROR(INDEX(\'风险登记表\'!$M$2:$M${risk_last_row},MATCH(B{row},\'风险登记表\'!$A$2:$A${risk_last_row},0)),"一般")')
        proc.cell(row, 10, f'=IF(OR(I{row}="高",I{row}="中高"),"增强","基础")')
        proc.cell(row, 15, f'=IF(J{row}="增强",8,4)')
        proc.cell(row, 24, f'=IF(\'项目参数\'!$B$2="计划",0,IF(OR(M{row}="",R{row}="已完成"),0,MAX(0,NETWORKDAYS(M{row}+1,\'项目参数\'!$B$3))))')
        for col in (12, 13, 14): proc.cell(row, col).number_format = "yyyy-mm-dd"
        proc.cell(row, 17).number_format = "0%"
    _style_header(proc, 1, 1, len(headers)); _style_body(proc, 2, proc.max_row, 1, len(headers))
    _set_widths(proc, [12, 14, 20, 42, 26, 58, 20, 26, 12, 14, 24, 14, 14, 14, 12, 12, 14, 12, 18, 28, 20, 28, 14, 14, 28, 14, 18])
    proc.freeze_panes = "D2"
    for col, values in [(18, '"未开始,进行中,已完成,受阻"'), (23, '"未提供,已提供,待补充"'), (26, '"未复核,复核中,已复核,退回修改"')]:
        from openpyxl.utils import get_column_letter
        letter = get_column_letter(col)
        dv = DataValidation(type="list", formula1=values)
        proc.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}{proc.max_row}")
    proc.conditional_formatting.add(f"R2:R{proc.max_row}", FormulaRule(formula=['$R2="已完成"'], fill=PatternFill("solid", fgColor="E2F0D9")))
    proc.conditional_formatting.add(f"R2:R{proc.max_row}", FormulaRule(formula=['$R2="受阻"'], fill=PatternFill("solid", fgColor="FDE9E7")))
    proc.conditional_formatting.add(f"W2:W{proc.max_row}", FormulaRule(formula=['$W2="已提供"'], fill=PatternFill("solid", fgColor="E2F0D9")))
    proc.conditional_formatting.add(f"Z2:Z{proc.max_row}", FormulaRule(formula=['$Z2="已复核"'], fill=PatternFill("solid", fgColor="E2F0D9")))
    _replace_table(proc, "AuditProcedures", f"A1:AA{proc.max_row}")

    _add_materiality_sheet(wb)
    _add_risk_sheet(wb)
    _add_controls_sheet(wb)
    _add_project_parameters(wb)
    _add_dashboard(wb, proc.max_row)

    target = OUTPUT_DIR / PLAN_OUTPUT_NAME
    target.parent.mkdir(parents=True, exist_ok=True)
    _enable_formula_recalculation(wb)
    wb.save(target)
    reviewer_conflicts = [row for row in range(2, proc.max_row + 1) if proc.cell(row, 7).value == proc.cell(row, 8).value]
    result = {
        "file": target.name,
        "sheet_names": wb.sheetnames,
        "sheet_count": len(wb.sheetnames),
        "procedure_count": proc.max_row - 1,
        "self_review_conflicts": reviewer_conflicts,
        "uses_networkdays": all(str(timeline.cell(row, 5).value).startswith("=NETWORKDAYS") for row in range(5, timeline.max_row + 1)),
        "project_mode": "计划",
        "project_status_date": "2023-01-10",
        "risk_mapping_count": proc.max_row - 1,
        "new_sheets": ["项目参数", "重要性与抽样", "风险登记表", "控制了解", "项目管理看板"],
    }
    return result


def build_all_workbooks() -> dict[str, object]:
    result = {"checklist": build_checklist_workbook(), "plan": build_audit_plan_workbook()}
    (OUTPUT_DIR / WORKBOOK_VERIFICATION_OUTPUT_NAME).write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    return result
