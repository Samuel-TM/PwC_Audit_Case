import json
import re
from datetime import date, datetime

import fitz
from openpyxl import load_workbook

from src.config import (
    CHART_DIR,
    CHECKLIST_OUTPUT_NAME,
    OUTPUT_DIR,
    PDF_VERIFICATION_OUTPUT_NAME,
    PLAN_OUTPUT_NAME,
    QA_OUTPUT_NAME,
    REPORT_OUTPUT_STEM,
    REPORT_TEX,
    RISK_OUTPUT_NAME,
    ROOT,
    SUMMARY_OUTPUT_NAME,
    WORKBOOK_VERIFICATION_OUTPUT_NAME,
)


def _pdf_text() -> str:
    pdf = fitz.open(OUTPUT_DIR / f"{REPORT_OUTPUT_STEM}.pdf")
    return "\n".join(page.get_text() for page in pdf)


def test_v4_charts_exist_and_are_projection_ready():
    from PIL import Image

    charts = sorted(CHART_DIR.glob("*.png"))
    assert len(charts) == 9
    for required in ["06_receivables_net_ratio.png", "07_inventory_composition_2022.png", "08_audit_risk_ranking.png", "09_audit_plan_gantt.png"]:
        assert (CHART_DIR / required).exists()
    for chart in charts:
        assert Image.open(chart).size == (1600, 900)


def test_v4_workbook_sheet_names_and_freeze_panes():
    checklist = load_workbook(OUTPUT_DIR / CHECKLIST_OUTPUT_NAME, read_only=False)
    assert checklist.sheetnames == ["总览", "通用资料", "收入", "应收账款", "存货", "采购与成本", "资金与借款", "其他重点项目", "其他报表项目", "费用税务与披露"]
    assert checklist["总览"].freeze_panes == "A6"
    assert checklist["收入"].freeze_panes == "A2"
    plan = load_workbook(OUTPUT_DIR / PLAN_OUTPUT_NAME, read_only=False)
    assert plan.sheetnames == ["项目管理看板", "项目参数", "总体时间表", "人员分工", "审计程序", "问题跟踪", "函证跟踪", "盘点安排", "重要性与抽样", "风险登记表", "控制了解"]
    assert plan["总体时间表"].freeze_panes == "A5"
    assert plan["审计程序"].freeze_panes == "D2"


def test_single_latex_source_pdf_layout_and_forbidden_text():
    tex = REPORT_TEX.read_text(encoding="utf-8")
    assert tex.count(r"\begin{frame}") == 15
    for macro in [r"\newcommand{\bodyfont}{\fontsize{10.8}{13.2}", r"\newcommand{\smallbody}{\fontsize{9.8}{12.0}", r"\newcommand{\tinybody}{\fontsize{8.5}{10.4}", r"\newcommand{\sourcefont}{\fontsize{7.0}{8.3}"]:
        assert macro in tex
    assert tex.count("assets/derived/PwC_logo_cropped.png") == 2
    assert "08_audit_risk_ranking.png" in tex
    assert "output/output_V4/charts_V4/" in tex
    assert not list(OUTPUT_DIR.glob("*.tex"))
    pdf = fitz.open(OUTPUT_DIR / f"{REPORT_OUTPUT_STEM}.pdf")
    assert pdf.page_count == 15
    assert all(abs(page.rect.width / page.rect.height - 16 / 9) < 0.01 for page in pdf)
    text = "\n".join(page.get_text() for page in pdf)
    for forbidden in ["待填写", ".md", "_V3", "回答时先", "台账保持待执行", "年末业绩压力", "地点分散"]:
        assert forbidden not in text
    for page_index in range(1, 12):
        assert "来源：" in pdf[page_index].get_text()


def test_v4_page_specific_content_and_chart_heights():
    tex = REPORT_TEX.read_text(encoding="utf-8")
    expected = {
        "01_revenue_net_profit.png": "4.40cm",
        "02_receivables_inventory.png": "5.00cm",
        "08_audit_risk_ranking.png": "4.70cm",
        "05_export_share.png": "3.45cm",
        "06_receivables_net_ratio.png": "3.30cm",
        "07_inventory_composition_2022.png": "3.45cm",
        "04_product_gross_margin.png": "3.45cm",
        "09_audit_plan_gantt.png": "5.00cm",
    }
    for chart, height in expected.items():
        assert re.search(rf"height={re.escape(height)}[^\n]+{re.escape(chart)}", tex)
    for phrase in ["业务背景与审计策略", "80.40", "72.92", "模拟项目时间轴", "计划阶段拟重点审计事项", "先确定基准，再设计样本", "Thank you"]:
        assert phrase in tex


def test_plan_governance_and_formula_driven_risk_mapping():
    plan = load_workbook(OUTPUT_DIR / PLAN_OUTPUT_NAME, data_only=False)
    procedures = plan["审计程序"]
    rows = list(procedures.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 30
    assert all(row[6] != row[7] for row in rows)
    assert all(str(row[8]).startswith("=IFERROR(INDEX('风险登记表'!") for row in rows)
    assert all(str(row[9]).startswith("=IF(OR(I") for row in rows)
    valid_ids = {plan["风险登记表"].cell(row, 1).value for row in range(2, 10)}
    assert all(row[1] in valid_ids for row in rows)
    mapped = {row[0]: row[1] for row in rows}
    assert mapped["P-01"] == mapped["OCA-01"] == "R-PREOCA"
    assert mapped["B-01"] == mapped["GC-01"] == "R-DEBTGC"


def test_roles_are_separate_and_no_self_review():
    plan = load_workbook(OUTPUT_DIR / PLAN_OUTPUT_NAME, data_only=False)
    roles = plan["人员分工"]
    role_names = [roles.cell(row, 1).value for row in range(2, roles.max_row + 1)]
    assert "项目合伙人" in role_names
    assert "质量复核人（如适用）" in role_names
    assert all("/" not in str(role) for role in role_names)
    assert all(plan["审计程序"].cell(row, 7).value != plan["审计程序"].cell(row, 8).value for row in range(2, plan["审计程序"].max_row + 1))


def test_project_mode_overdue_and_dashboard_formulas():
    plan = load_workbook(OUTPUT_DIR / PLAN_OUTPUT_NAME, data_only=False)
    params = plan["项目参数"]
    assert params["B2"].value == "计划"
    assert params["B3"].value in {date(2023, 1, 10), datetime(2023, 1, 10)}
    procedures = plan["审计程序"]
    for row in range(2, procedures.max_row + 1):
        formula = procedures.cell(row, 24).value
        assert "'项目参数'!$B$2=\"计划\"" in formula
        assert "'项目参数'!$B$3" in formula
        assert "TODAY()" not in formula
    dashboard = plan["项目管理看板"]
    assert "'审计程序'!$R$2:$R$" in dashboard["B4"].value
    assert "COUNTIFS" in dashboard["E4"].value
    assert dashboard["B6"].value == f"=IFERROR('[{CHECKLIST_OUTPUT_NAME}]总览'!$I$15,0)"


def test_materiality_risk_and_control_templates():
    plan = load_workbook(OUTPUT_DIR / PLAN_OUTPUT_NAME, data_only=False)
    materiality = plan["重要性与抽样"]
    assert materiality["B5"].value.startswith("=IF(")
    assert "候选基准" in materiality["B2"].value
    risk = plan["风险登记表"]
    assert risk["A1"].value == "风险 ID"
    assert risk["I2"].value == "=AVERAGE(C2:H2)"
    assert risk["M2"].value == '=IF(K2="",J2,K2)'
    assert risk["A9"].value == "R-BASE"
    controls = plan["控制了解"]
    assert controls.max_row == 7
    assert {controls.cell(row, 1).value for row in range(2, 8)} == {"销售与收款", "采购与付款", "生产与存货", "资金", "财务结账", "主数据和系统权限"}


def test_checklist_priority_policy_and_historical_date_logic():
    checklist = load_workbook(OUTPUT_DIR / CHECKLIST_OUTPUT_NAME, data_only=False)
    overview = checklist["总览"]
    assert overview.tables["ChecklistOverview"].ref == "A5:J15"
    assert overview["M2"].value == "计划"
    assert overview["B15"].value == "=SUM(B6:B14)"
    assert overview["I6"].value == "=IF(B6=0,0,F6/B6)"
    counts = {"高": 0, "中": 0, "低": 0}
    for sheet_name in checklist.sheetnames[1:]:
        ws = checklist[sheet_name]
        assert ws["L1"].value == "优先级依据/备注"
        assert any(str(item.sqref).startswith("H2:H") for item in ws.data_validations.dataValidation)
        assert any(str(item.sqref).startswith("K2:K") for item in ws.data_validations.dataValidation)
        rules = ws.conditional_formatting._cf_rules.values()
        formulas = [formula for entries in rules for entry in entries for formula in (entry.formula or [])]
        assert any("'总览'!$M$2=\"执行\"" in formula for formula in formulas)
        assert all("TODAY()" not in formula for formula in formulas)
        for row in range(2, ws.max_row + 1):
            priority = ws.cell(row, 8).value
            counts[priority] += 1
            assert ws.cell(row, 12).value
            planned = ws.cell(row, 9).value
            planned_date = planned.date() if isinstance(planned, datetime) else planned
            if priority == "高":
                assert planned_date <= date(2023, 1, 6)
            elif priority == "低":
                assert planned_date >= date(2023, 1, 20)
    assert sum(counts.values()) == 101
    assert counts["高"] / 101 < 0.5


def test_plan_validations_and_conditional_formats():
    plan = load_workbook(OUTPUT_DIR / PLAN_OUTPUT_NAME, data_only=False)
    procedures = plan["审计程序"]
    validation_ranges = {str(item.sqref) for item in procedures.data_validations.dataValidation}
    assert len(validation_ranges) == 3
    assert any(value.startswith("R2:R") for value in validation_ranges)
    assert any(value.startswith("W2:W") for value in validation_ranges)
    assert any(value.startswith("Z2:Z") for value in validation_ranges)
    assert sum(len(rules) for rules in procedures.conditional_formatting._cf_rules.values()) >= 4


def test_qa_numbering_page_indexes_and_required_answers():
    text = (OUTPUT_DIR / QA_OUTPUT_NAME).read_text(encoding="utf-8")
    assert text.count("**30 秒回答：**") == 32
    assert text.count("对应 PDF 第") == 32
    for phrase in ["第四季度收入占比下降", "旧版 78 项高优先级为什么调整", "如存在异地存货", "质量复核人（如适用）"]:
        assert phrase in text
    assert "第四季度收入占比下降，为什么仍做截止测试？（对应 PDF 第 6 页）" in text
    assert "旧版 78 项高优先级为什么调整？（对应 PDF 第 10 页）" in text


def test_cross_file_key_numbers_and_verification_outputs():
    pdf_text = _pdf_text()
    summary = (OUTPUT_DIR / SUMMARY_OUTPUT_NAME).read_text(encoding="utf-8")
    for value in ["1.22%", "34.35%", "21.38%", "23.61%", "22.53%", "28.90%", "648.33", "4,109.52", "4,504.14", "101", "30"]:
        assert value in pdf_text or value in summary
    workbook_report = json.loads((OUTPUT_DIR / WORKBOOK_VERIFICATION_OUTPUT_NAME).read_text(encoding="utf-8"))
    assert workbook_report["checklist"]["total_requests"] == 101
    assert workbook_report["checklist"]["high_priority_ratio"] < 0.5
    assert workbook_report["plan"]["procedure_count"] == 30
    assert workbook_report["plan"]["self_review_conflicts"] == []
    priorities = workbook_report["checklist"]["priority_counts"]
    assert sum(priorities.values()) == 101
    for count in priorities.values():
        assert str(count) in pdf_text and str(count) in summary
    risks = json.loads((OUTPUT_DIR / RISK_OUTPUT_NAME).read_text(encoding="utf-8"))
    assert {item["risk_id"] for item in risks} >= {"R-PREOCA", "R-DEBTGC", "R-EQUITY"}
    pdf_report = json.loads((OUTPUT_DIR / PDF_VERIFICATION_OUTPUT_NAME).read_text(encoding="utf-8"))
    assert pdf_report["status"] == "PASS"
    assert pdf_report["checks"]["data_pages_have_sources"] is True


def test_build_source_has_no_pptx_or_tex_delivery():
    build_source = (ROOT / "scripts" / "build_all.py").read_text(encoding="utf-8")
    export_source = (ROOT / "scripts" / "export_pdf.py").read_text(encoding="utf-8")
    assert "未修改 PPTX" not in build_source
    assert "A公司审计案例汇报.pptx" not in build_source
    assert 'f"{REPORT_OUTPUT_STEM}.tex"' not in build_source
    assert "copy2(REPORT_TEX" not in export_source
